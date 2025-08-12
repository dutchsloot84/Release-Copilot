from __future__ import annotations

from pathlib import Path
from typing import Dict, List
import csv

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def _freeze_and_filter(ws: Worksheet) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def _auto_fit(ws: Worksheet) -> None:
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2


def _add_csv_sheet(wb: Workbook, sheet_title: str, csv_path: Path):
    ws = wb.create_sheet(title=sheet_title[:31])
    rows = []
    if csv_path.exists():
        with csv_path.open("r", encoding="utf-8", newline="") as f:
            reader = csv.reader(f)
            for row in reader:
                rows.append(row)
    if rows:
        for row in rows:
            ws.append(row)
    else:
        ws.append(["No data"])
    _freeze_and_filter(ws)
    _auto_fit(ws)


def build_reports(summary_rows: List[Dict], output_dir: Path, repo_csv_map: Dict[str, Path], base_name: str = "release_audit") -> None:
    # Markdown
    md_lines = ["| Project | Repo | Branch | Count | Source |", "|---|---|---|---|---|"]
    for r in summary_rows:
        md_lines.append(
            f"| {r.get('project','')} | {r.get('repo','')} | {r.get('branch','')} | {r.get('count',0)} | {r.get('source','')} |")
    miss_p = output_dir / "missing_in_repo.csv"
    orph_p = output_dir / "orphan_commits.csv"
    miss_ct = sum(1 for _ in open(miss_p, "r", encoding="utf-8")) - 1 if miss_p.exists() else 0
    orph_ct = sum(1 for _ in open(orph_p, "r", encoding="utf-8")) - 1 if orph_p.exists() else 0
    md_lines.append("")
    md_lines.append(f"**Jira comparison:** {miss_ct} issue(s) missing in repo · {orph_ct} orphan commit(s).")
    if miss_p.exists():
        md_lines.append(f"- [Missing in repo CSV]({miss_p.as_posix()})")
    if orph_p.exists():
        md_lines.append(f"- [Orphan commits CSV]({orph_p.as_posix()})")
    md_path = output_dir / f"{base_name}.md"
    md_path.write_text("\n".join(md_lines), encoding="utf-8")

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["project", "repo", "branch", "count", "csv_path", "source"])
    for r in summary_rows:
        ws.append([
            r.get("project", ""),
            r.get("repo", ""),
            r.get("branch", ""),
            r.get("count", 0),
            r.get("csv_path", ""),
            r.get("source", ""),
        ])
    _freeze_and_filter(ws)
    _auto_fit(ws)
    # per-repo sheets
    for repo, csv_path in repo_csv_map.items():
        _add_csv_sheet(wb, repo[:31], csv_path)
    if miss_p.exists():
        _add_csv_sheet(wb, "MissingInRepo", miss_p)
    if orph_p.exists():
        _add_csv_sheet(wb, "OrphanCommits", orph_p)
    xlsx_path = output_dir / f"{base_name}.xlsx"
    wb.save(xlsx_path)
