from pathlib import Path
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def _autowidth(ws: Worksheet) -> None:
    for column_cells in ws.columns:
        length = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2


def write_excel_audit(jira_list: List[Dict], commit_list: List[Dict], matches: List[Dict],
                       missing_in_git: List[Dict], commits_without_story: List[Dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Matches'
    headers = ['Jira Key', 'Summary', 'Commit', 'Author']
    ws.append(headers)
    for m in matches:
        ws.append([m.get('key'), m.get('summary'), m.get('commit'), m.get('author')])
    ws.auto_filter.ref = f"A1:D{ws.max_row}"
    ws.freeze_panes = 'A2'
    _autowidth(ws)

    ws2 = wb.create_sheet('MissingInGit')
    ws2.append(['Jira Key', 'Summary'])
    for j in missing_in_git:
        ws2.append([j.get('key'), j.get('summary')])
    ws2.auto_filter.ref = f"A1:B{ws2.max_row}"
    ws2.freeze_panes = 'A2'
    _autowidth(ws2)

    ws3 = wb.create_sheet('CommitsWithoutStory')
    ws3.append(['Commit', 'Author'])
    for c in commits_without_story:
        ws3.append([c.get('id'), c.get('author')])
    ws3.auto_filter.ref = f"A1:B{ws3.max_row}"
    ws3.freeze_panes = 'A2'
    _autowidth(ws3)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_markdown_report(summary: str, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(summary)
